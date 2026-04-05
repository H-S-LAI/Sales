import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, timedelta
import io

# ════════════════════════════════════════════════════════════════
# 常數定義
# ════════════════════════════════════════════════════════════════

G1_STORES = ['彰草店', '金美店', '日華店']
G2_STORES = ['北屯店', '向上店', '五權店', '太平店', '大雅店', '漢口店']
G3_STORES = ['金馬店', '正德店', '大埔店', '三民店', '線東店', '彰美店',
             '過溝店', '彰鹿店', '泰和店', '精誠店', '秀二店', '花壇店', '華山店']

G1_ITEMS  = ['特幼', '幼大口', '多粒', '多大口', '幼菁', '雙子星']
G2_ITEMS  = ['特幼', '多菁',   '幼大口', '多粒', '多大口', '幼菁', '雙子星']
G3_ITEMS  = ['特幼', '普通',   '幼大口', '多粒', '多大口', '幼菁', '雙子星']
GRAND_ITEMS = ['特幼', '幼大口', '多粒', '多大口', '幼菁', '雙子星']

# 每個品名對應 Excel 欄位 (品名欄, 售量欄) — 1-indexed
ITEM_COLS = {
    '特幼':   (2,  3),
    '多菁':   (4,  5),
    '普通':   (4,  5),
    '幼大口': (6,  7),
    '多粒':   (8,  9),
    '多大口': (10, 11),
    '幼菁':   (12, 13),
    '雙子星': (14, 15),
}

# 預設粒換算（若無前日 sheet 可讀則使用）
FALLBACK_RATES = {
    'g1': {'特幼': 5, '幼大口': 5,  '多粒': 9,  '多大口': 9,  '幼菁': 6, '雙子星': 14},
    'g2': {'特幼': 6, '多菁': 10,   '幼大口': 6, '多粒': 14, '多大口': 15, '幼菁': 9, '雙子星': 14},
    'g3': {'特幼': 6, '普通': 8,    '幼大口': 6, '多粒': 14, '多大口': 15, '幼菁': 9, '雙子星': 14},
}

# ════════════════════════════════════════════════════════════════
# 樣式定義
# ════════════════════════════════════════════════════════════════

_s  = Side(style='thin')
BRD = Border(left=_s, right=_s, top=_s, bottom=_s)
AC  = Alignment(horizontal='center', vertical='center')
AL  = Alignment(horizontal='left',   vertical='center')
FHD = PatternFill('solid', fgColor='D9E1F2')  # 標頭淺藍底

def F(sz=12, bold=False, c="000000"):
    return Font(name='新細明體', size=sz, bold=bold, color=c)

F_N = F()                       # 黑色一般
F_B = F(bold=True)              # 黑色粗體（店名）
F_T = F(14)                     # 標題 14pt
F_G = F(bold=True, c="008000")  # 🟢 綠色 — 粒換算常數
F_R = F(bold=True, c="FF0000")  # 🔴 紅色 — 銷售包數合計
F_U = F(bold=True, c="0000FF")  # 🔵 藍色 — 粒數（包數×換算）


def sc(ws, r, c, v=None, f=None, a=AC, b=BRD, fi=None):
    """快速設定儲存格"""
    cl = ws.cell(r, c)
    if v is not None: cl.value = v
    if f:  cl.font = f
    if a:  cl.alignment = a
    if b:  cl.border = b
    if fi: cl.fill = fi
    return cl


# ════════════════════════════════════════════════════════════════
# 讀取當日售量
# ════════════════════════════════════════════════════════════════

def load_sales(file_bytes):
    df = pd.read_excel(io.BytesIO(file_bytes), engine='openpyxl')
    cs = ci = cq = None
    for col in df.columns:
        s = str(col)
        if '店名' in s: cs = col
        if '品名' in s: ci = col
        if '售量' in s: cq = col
    if not all([cs, ci, cq]):
        return None, "⚠️ 找不到必要欄位（店名 / 品名 / 售量），請確認 Raw_data 格式"
    df = df[[cs, ci, cq]].copy()
    df.columns = ['店名', '品名', '售量']
    df['店名'] = df['店名'].astype(str).str.strip()
    df['品名'] = df['品名'].astype(str).str.strip()
    df = df[df['店名'].ne('') & df['店名'].ne('nan') & df['品名'].ne('nan')]
    df['售量'] = pd.to_numeric(df['售量'], errors='coerce').fillna(0).astype(int)
    result = {}
    for _, row in df.iterrows():
        s, i, q = row['店名'], row['品名'], row['售量']
        if s and i:
            result.setdefault(s, {})[i] = q
    return result, None


# ════════════════════════════════════════════════════════════════
# 從前一天 sheet 讀粒換算
# ════════════════════════════════════════════════════════════════

def read_prev_rates(file_bytes, fname, target_date):
    """讀取前一天 sheet 的 🟢 綠色粒換算數字；讀不到則用 FALLBACK"""
    rates = {g: dict(d) for g, d in FALLBACK_RATES.items()}
    if not file_bytes:
        return rates, "（無累計檔，使用預設值）"

    prev = target_date - timedelta(days=1)
    sn   = f"{prev.month}-{prev.day}"
    gmap = [('g1', G1_ITEMS), ('g2', G2_ITEMS), ('g3', G3_ITEMS)]

    try:
        if fname.lower().endswith('.xls'):
            import xlrd
            wb_x  = xlrd.open_workbook(file_contents=file_bytes)
            names = wb_x.sheet_names()
            ws_x  = wb_x.sheet_by_name(sn) if sn in names else wb_x.sheet_by_index(-1)
            used  = ws_x.name
            rows_idx = [i for i in range(ws_x.nrows) if ws_x.cell_value(i, 0) == '銷售包數']
            for gi, (gk, items) in enumerate(gmap):
                if gi >= len(rows_idx): break
                ri = rows_idx[gi]
                for item in items:
                    nc, _ = ITEM_COLS[item]
                    v = ws_x.cell_value(ri, nc - 1)
                    if isinstance(v, (int, float)) and v > 0:
                        rates[gk][item] = int(v)
            return rates, f"（從 .xls「{used}」讀取）"
        else:
            wb   = load_workbook(io.BytesIO(file_bytes), data_only=True)
            if sn in wb.sheetnames:
                ws = wb[sn]
                used = sn
            else:
                ws   = wb[wb.sheetnames[-1]]
                used = wb.sheetnames[-1]
            prows = [r for r in ws.iter_rows() if r[0].value == '銷售包數']
            for gi, (gk, items) in enumerate(gmap):
                if gi >= len(prows): break
                row = prows[gi]
                for item in items:
                    nc, _ = ITEM_COLS[item]
                    v = row[nc - 1].value
                    if isinstance(v, (int, float)) and v > 0:
                        rates[gk][item] = int(v)
            return rates, f"（從「{used}」讀取）"
    except Exception as e:
        return rates, f"（讀取失敗，使用預設值：{e}）"


# ════════════════════════════════════════════════════════════════
# 生成工作表
# ════════════════════════════════════════════════════════════════

def build_sheet(ws, sales, rates, d):
    # 欄寬
    # 欄寬：完全對應原始 .xls（A=9.62，B~O=8.5）
    ws.column_dimensions["A"].width = 9.62
    for ci in range(2, 16):
        ws.column_dimensions[chr(64 + ci)].width = 8.5

    # 第 1 列：標題
    tw = d.year - 1911
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
    sc(ws, 1, 1, f"{tw}.{d.month}.{d.day}檳榔銷售統計", F_T, AL, None)
    ws.row_dimensions[1].height = 20.2

    HDR = ['店名','品名','售量','品名','售量','品名','售量','品名','售量',
           '品名','售量','品名','售量','品名','售量']

    def write_header(r):
        for ci, h in enumerate(HDR, 1):
            sc(ws, r, ci, h, F(12), AC, BRD, FHD)
        ws.row_dimensions[r].height = 19.5

    write_header(2)

    def write_group(stores, items, r0, gk):
        r  = r0
        gr = rates[gk]

        # ── 各店資料列 ──
        for store in stores:
            ws.row_dimensions[r].height = 19.5
            sd = sales.get(store, {})
            sc(ws, r, 1, store, F_B)
            # 先把全部 14 欄填空（帶框線）
            for ci in range(2, 16):
                sc(ws, r, ci, f=F_N)
            # 再填入此群組有的品名 & 售量
            for item in items:
                nc, qc = ITEM_COLS[item]
                sc(ws, r, nc, item, F_N)
                sc(ws, r, qc, sd.get(item, 0) or None, F_N)
            r += 1

        # 各品名當日合計
        tots = {it: sum(sales.get(s, {}).get(it, 0) for s in stores) for it in items}

        # ── 銷售包數 列（紅字合計 / 綠字換算） ──
        ws.row_dimensions[r].height = 19.5
        sc(ws, r, 1, '銷售包數', F_R)
        for ci in range(2, 16):
            sc(ws, r, ci, f=F_R)
        for item in items:
            nc, qc = ITEM_COLS[item]
            sc(ws, r, nc, gr.get(item, 0), F_G)           # 🟢 粒換算
            sc(ws, r, qc, tots[item] or None, F_R)         # 🔴 包數合計
        r += 1

        # ── 銷售粒數 列（藍字：包數 × 換算） ──
        ws.row_dimensions[r].height = 19.5
        sc(ws, r, 1, '銷售粒數', F_U)
        for ci in range(2, 16):
            sc(ws, r, ci, f=F_U)
        for item in items:
            _, qc = ITEM_COLS[item]
            粒 = tots[item] * gr.get(item, 0)
            sc(ws, r, qc, 粒 or None, F_U)                 # 🔵 粒數
        r += 1

        return r, tots

    r, t1 = write_group(G1_STORES, G1_ITEMS, 3,  'g1')
    r, t2 = write_group(G2_STORES, G2_ITEMS, r,  'g2')
    write_header(r); r += 1
    r, t3 = write_group(G3_STORES, G3_ITEMS, r,  'g3')

    # ── 第 32 列：總合計 ──
    ws.row_dimensions[r].height = 19.5
    all_s   = G1_STORES + G2_STORES + G3_STORES
    tot_pkg = sum(sum(sales.get(s, {}).values()) for s in all_s)
    sc(ws, r, 1, tot_pkg, F_R)                             # 🔴 全部包數
    for ci in range(2, 16):
        sc(ws, r, ci, f=F_N)
    for item in GRAND_ITEMS:
        _, qc = ITEM_COLS[item]
        total = t1.get(item, 0) + t2.get(item, 0) + t3.get(item, 0)
        sc(ws, r, qc, total or None, F_N)                  # ⬛ 各共同品名合計


# ════════════════════════════════════════════════════════════════
# Streamlit UI
# ════════════════════════════════════════════════════════════════

st.set_page_config(page_title="檳榔銷售統計產生器", layout="wide")
st.title("🌿 檳榔銷售統計自動化系統")

# ── 上傳區 ──
c1, c2 = st.columns(2)
with c1:
    f_raw = st.file_uploader("① 上傳從檳榔管理系統匯出之原始檔", type=['xlsx'])
with c2:
    f_cum = st.file_uploader(
        "② 上傳現有累計檔（.xls 或 .xlsx，用於讀取前日粒換算）",
        type=['xls', 'xlsx']
    )

# ── 日期選擇（預設昨天）──
default_d   = (datetime.now() - timedelta(days=1)).date()
rd          = st.date_input("📅 報表日期（預設昨天，可修改）", value=default_d)
report_date = datetime(rd.year, rd.month, rd.day)

# ── 讀入售量 ──
sales, err = {}, None
if f_raw:
    sales, err = load_sales(f_raw.getvalue())
    if err:    st.error(err)
    elif sales: st.caption(f"✅ 成功讀入 {len(sales)} 家店資料")

# ── 讀入粒換算 ──
cum_bytes = f_cum.getvalue() if f_cum else None
cum_name  = f_cum.name      if f_cum else ""
rates, rate_msg = read_prev_rates(cum_bytes, cum_name, report_date)

# ── 粒換算設定 UI（表格形式）──
st.markdown("---")
st.markdown(f"### 🟢 粒換算設定　{rate_msg}")
st.caption("每個品名「一包等於幾粒」，自動從前日帶入，有需要可手動調整。")

# 欄位：店區 | 特幼 | 多菁 | 普通 | 幼大口 | 多粒 | 多大口 | 幼菁 | 雙子星
# 日紅：多菁=— 普通=—  台中：普通=—  彰化：多菁=—
COL_HDRS  = ['店區', '特幼', '多菁', '普通', '幼大口', '多粒', '多大口', '幼菁', '雙子星']
ITEM_ROWS = {
    'g1': {'特幼':'特幼', '多菁':None,   '普通':None,   '幼大口':'幼大口', '多粒':'多粒', '多大口':'多大口', '幼菁':'幼菁', '雙子星':'雙子星'},
    'g2': {'特幼':'特幼', '多菁':'多菁', '普通':None,   '幼大口':'幼大口', '多粒':'多粒', '多大口':'多大口', '幼菁':'幼菁', '雙子星':'雙子星'},
    'g3': {'特幼':'特幼', '多菁':None,   '普通':'普通', '幼大口':'幼大口', '多粒':'多粒', '多大口':'多大口', '幼菁':'幼菁', '雙子星':'雙子星'},
}
GROUPS = [('g1', '日紅'), ('g2', '台中'), ('g3', '彰化')]
col_w  = [0.9, 1, 1, 1, 1, 1, 1, 1, 1]

# 表頭
hdr = st.columns(col_w)
for i, h in enumerate(COL_HDRS):
    hdr[i].markdown(f"**{h}**")
st.markdown('<hr style="margin:2px 0 6px 0">', unsafe_allow_html=True)

# 資料列
for gk, label in GROUPS:
    row   = st.columns(col_w)
    items = ITEM_ROWS[gk]
    row[0].markdown(f"<div style='padding-top:10px'><b>{label}</b></div>", unsafe_allow_html=True)
    for ci, key in enumerate(COL_HDRS[1:], 1):
        actual = items[key]
        if actual is None:
            row[ci].markdown("<div style='text-align:center;color:#bbb;padding-top:10px;font-size:18px'>—</div>",
                             unsafe_allow_html=True)
        else:
            rates[gk][actual] = row[ci].number_input(
                key, min_value=0, max_value=99,
                value=int(rates[gk].get(actual, 0)),
                key=f"rate_{gk}_{actual}",
                step=1, label_visibility="collapsed"
            )

# ── 生成按鈕 ──
st.markdown("---")
if st.button("🚀 生成報表", type="primary", disabled=(not f_raw or bool(err))):
    if not sales:
        st.error("Raw_data 無法讀取，請確認檔案格式")
        st.stop()

    # 建立 / 載入累計活頁簿
    if cum_bytes and cum_name.lower().endswith('.xlsx'):
        wb = load_workbook(io.BytesIO(cum_bytes))
    else:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    # 新增工作表（position=0 → 放最前面；若已存在則先刪除）
    sn = f"{report_date.month}-{report_date.day}"
    if sn in wb.sheetnames:
        del wb[sn]
    ws = wb.create_sheet(sn, 0)

    build_sheet(ws, sales, rates, report_date)

    out = io.BytesIO()
    wb.save(out)

    tw    = report_date.year - 1911
    fname = f"檳榔銷售統計_{tw}年{report_date.month}月{report_date.day}日.xlsx"
    st.success(f"✅ 工作表「{sn}」生成完成，已放在最前面！")
    st.download_button(
        "💾 下載 Excel",
        data=out.getvalue(),
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
